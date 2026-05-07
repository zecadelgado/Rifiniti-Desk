"""
Controller para Importação em Massa de Patrimônios

Versão: 2.4
Data: 19/11/2025
"""

import os
from typing import Dict, List, Optional

from PySide6.QtCore import Qt, QThread
from PySide6.QtWidgets import (
    QDialog,
    QFileDialog,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QGroupBox
)

from backend.database.database_manager import DatabaseManager
from backend.services.import_patrimonio import PatrimonioImporter, ImportResult
from backend.services.audit_helper import registrar_auditoria


class ImportThread(QThread):
    """Thread para executar importação em background"""
    
    def __init__(self, importer: PatrimonioImporter, linhas):
        super().__init__()
        self.importer = importer
        self.linhas = linhas
        self.result: Optional[ImportResult] = None
    
    def run(self):
        self.result = self.importer.importar(self.linhas)


class ImportPatrimonioDialog(QDialog):
    """Diálogo para importação em massa de patrimônios"""
    
    def __init__(self, db_manager: DatabaseManager, current_user: Optional[Dict] = None, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.current_user = current_user
        self.importer = PatrimonioImporter(db_manager)
        self.dados_validados = None
        self.linhas_validas = None
        
        self.setWindowTitle("Importação em Massa de Patrimônios")
        self.setMinimumSize(900, 700)
        
        self._setup_ui()
        self._connect_signals()
    
    def _setup_ui(self):
        """Configura a interface"""
        layout = QVBoxLayout(self)
        
        # Grupo 1: Seleção de arquivo
        group_arquivo = QGroupBox("1. Selecionar Arquivo")
        layout_arquivo = QVBoxLayout()
        
        # Botões de arquivo
        layout_btns_arquivo = QHBoxLayout()
        
        self.btn_template = QPushButton("📥 Baixar Template")
        self.btn_template.setToolTip("Baixar planilha modelo para preenchimento")
        layout_btns_arquivo.addWidget(self.btn_template)
        
        self.btn_selecionar = QPushButton("📂 Selecionar Planilha")
        self.btn_selecionar.setToolTip("Selecionar arquivo Excel (.xlsx) ou CSV")
        layout_btns_arquivo.addWidget(self.btn_selecionar)
        
        layout_arquivo.addLayout(layout_btns_arquivo)
        
        self.lbl_arquivo = QLabel("Nenhum arquivo selecionado")
        self.lbl_arquivo.setStyleSheet("color: gray; font-style: italic;")
        layout_arquivo.addWidget(self.lbl_arquivo)
        
        group_arquivo.setLayout(layout_arquivo)
        layout.addWidget(group_arquivo)
        
        # Grupo 2: Validação
        group_validacao = QGroupBox("2. Validar Dados")
        layout_validacao = QVBoxLayout()
        
        self.btn_validar = QPushButton("✓ Validar Planilha")
        self.btn_validar.setEnabled(False)
        self.btn_validar.setToolTip("Verificar se os dados estão corretos")
        layout_validacao.addWidget(self.btn_validar)
        
        self.lbl_validacao = QLabel("Aguardando seleção de arquivo...")
        self.lbl_validacao.setStyleSheet("color: gray;")
        layout_validacao.addWidget(self.lbl_validacao)
        
        # Tabela de preview
        self.table_preview = QTableWidget()
        self.table_preview.setMaximumHeight(200)
        self.table_preview.setVisible(False)
        layout_validacao.addWidget(self.table_preview)
        
        # Área de erros
        self.txt_erros = QTextEdit()
        self.txt_erros.setReadOnly(True)
        self.txt_erros.setMaximumHeight(150)
        self.txt_erros.setVisible(False)
        self.txt_erros.setStyleSheet("background-color: #fff3cd; color: #856404;")
        layout_validacao.addWidget(self.txt_erros)
        
        group_validacao.setLayout(layout_validacao)
        layout.addWidget(group_validacao)
        
        # Grupo 3: Importação
        group_importacao = QGroupBox("3. Importar")
        layout_importacao = QVBoxLayout()
        
        self.btn_importar = QPushButton("⬆ Iniciar Importação")
        self.btn_importar.setEnabled(False)
        self.btn_importar.setToolTip("Importar patrimônios para o banco de dados")
        self.btn_importar.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 10px;")
        layout_importacao.addWidget(self.btn_importar)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout_importacao.addWidget(self.progress_bar)
        
        self.lbl_status = QLabel("")
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_importacao.addWidget(self.lbl_status)
        
        group_importacao.setLayout(layout_importacao)
        layout.addWidget(group_importacao)
        
        # Botão fechar
        self.btn_fechar = QPushButton("Fechar")
        layout.addWidget(self.btn_fechar)
    
    def _connect_signals(self):
        """Conecta sinais"""
        self.btn_template.clicked.connect(self._baixar_template)
        self.btn_selecionar.clicked.connect(self._selecionar_arquivo)
        self.btn_validar.clicked.connect(self._validar_dados)
        self.btn_importar.clicked.connect(self._importar)
        self.btn_fechar.clicked.connect(self.close)
        
        # Sinais do importer
        self.importer.progress_updated.connect(self._atualizar_progresso)
        self.importer.status_updated.connect(self._atualizar_status)
    
    def _baixar_template(self):
        """Baixa template de planilha"""
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Template",
            "template_importacao_patrimonios.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filepath:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Patrimônios"
            
            # Cabeçalhos
            headers = [
                "nome",
                "descricao",
                "numero_serie",
                "data_aquisicao",
                "valor_compra",
                "quantidade",
                "numero_nota",
                "estado_conservacao",
                "categoria",
                "fornecedor_nome",
                "fornecedor_cnpj",
                "fornecedor_telefone",
                "fornecedor_email",
                "fornecedor_inscricao",
                "fornecedor_contato",
                "fornecedor_observacoes",
                "setor_local",
                "status"
            ]
            
            # Estilo do cabeçalho
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Linha de exemplo
            exemplo = [
                "Notebook Dell Latitude 5420",  # nome
                "Notebook corporativo com 16GB RAM",  # descricao
                "SN123456789",  # numero_serie
                "15/11/2025",  # data_aquisicao
                "3500.00",  # valor_compra
                "1",  # quantidade
                "12345",  # numero_nota
                "novo",  # estado_conservacao
                "Informática",  # categoria
                "Dell Computadores",  # fornecedor_nome
                "12.345.678/0001-90",  # fornecedor_cnpj
                "(11) 98765-4321",  # fornecedor_telefone
                "vendas@dell.com",  # fornecedor_email
                "",  # fornecedor_inscricao
                "João Silva",  # fornecedor_contato
                "",  # fornecedor_observacoes
                "TI - Sala 101",  # setor_local
                "ativo"  # status
            ]
            
            for col_idx, value in enumerate(exemplo, start=1):
                ws.cell(row=2, column=col_idx, value=value)
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Salvar
            wb.save(filepath)
            
            QMessageBox.information(
                self,
                "Template Criado",
                f"Template salvo com sucesso!\n\n{filepath}\n\nPreencha a planilha e importe."
            )
        
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erro",
                f"Erro ao criar template:\n{str(e)}"
            )
    
    def _selecionar_arquivo(self):
        """Seleciona arquivo para importação"""
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Planilha",
            "",
            "Planilhas (*.xlsx *.xls *.csv)"
        )
        
        if not filepath:
            return
        
        # Validar arquivo
        sucesso, msg = self.importer.validar_arquivo(filepath)
        
        if not sucesso:
            QMessageBox.warning(self, "Arquivo Inválido", msg)
            return
        
        # Ler planilha
        sucesso, dados, msg = self.importer.ler_planilha(filepath)
        
        if not sucesso:
            QMessageBox.critical(self, "Erro ao Ler Arquivo", msg)
            return
        
        # Armazenar dados
        self.dados_validados = dados
        self.linhas_validas = None
        
        # Atualizar UI
        self.lbl_arquivo.setText(f"📄 {os.path.basename(filepath)} ({len(dados)} linhas)")
        self.lbl_arquivo.setStyleSheet("color: green; font-weight: bold;")
        
        self.btn_validar.setEnabled(True)
        self.btn_importar.setEnabled(False)
        
        self.lbl_validacao.setText(f"Arquivo carregado. Clique em 'Validar Planilha' para continuar.")
        self.lbl_validacao.setStyleSheet("color: blue;")
        
        self.table_preview.setVisible(False)
        self.txt_erros.setVisible(False)
    
    def _validar_dados(self):
        """Valida os dados da planilha"""
        if not self.dados_validados:
            return
        
        self.lbl_validacao.setText("Validando dados...")
        self.lbl_validacao.setStyleSheet("color: blue;")
        
        # Validar
        sucesso, linhas_validas, erros = self.importer.validar_dados(self.dados_validados)
        
        if erros:
            # Mostrar erros
            self.txt_erros.setPlainText("\n".join(erros))
            self.txt_erros.setVisible(True)
            
            self.lbl_validacao.setText(f"❌ {len(erros)} erro(s) encontrado(s). Corrija e tente novamente.")
            self.lbl_validacao.setStyleSheet("color: red; font-weight: bold;")
            
            self.btn_importar.setEnabled(False)
        
        else:
            # Sucesso
            self.linhas_validas = linhas_validas
            self.txt_erros.setVisible(False)
            
            # Mostrar preview
            self._mostrar_preview(linhas_validas[:10])  # Primeiras 10 linhas
            
            self.lbl_validacao.setText(
                f"✓ Validação concluída! {len(linhas_validas)} patrimônio(s) pronto(s) para importar."
            )
            self.lbl_validacao.setStyleSheet("color: green; font-weight: bold;")
            
            self.btn_importar.setEnabled(True)
    
    def _mostrar_preview(self, linhas):
        """Mostra preview dos dados validados"""
        if not linhas:
            return
        
        self.table_preview.setVisible(True)
        self.table_preview.setRowCount(len(linhas))
        self.table_preview.setColumnCount(6)
        self.table_preview.setHorizontalHeaderLabels([
            "Nome", "Categoria", "Fornecedor", "Setor", "Valor", "Data"
        ])
        
        for row_idx, linha in enumerate(linhas):
            self.table_preview.setItem(row_idx, 0, QTableWidgetItem(linha.nome))
            self.table_preview.setItem(row_idx, 1, QTableWidgetItem(linha.categoria))
            self.table_preview.setItem(row_idx, 2, QTableWidgetItem(linha.fornecedor_nome))
            self.table_preview.setItem(row_idx, 3, QTableWidgetItem(linha.setor_local))
            self.table_preview.setItem(row_idx, 4, QTableWidgetItem(f"R$ {linha.valor_compra:.2f}"))
            self.table_preview.setItem(row_idx, 5, QTableWidgetItem(linha.data_aquisicao.strftime("%d/%m/%Y")))
        
        self.table_preview.resizeColumnsToContents()
    
    def _importar(self):
        """Inicia a importação"""
        if not self.linhas_validas:
            return
        
        # Confirmar
        total = len(self.linhas_validas)
        reply = QMessageBox.question(
            self,
            "Confirmar Importação",
            f"Deseja importar {total} patrimônio(s)?\n\n"
            "Fornecedores, setores e categorias serão criados automaticamente se não existirem.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Desabilitar botões
        self.btn_selecionar.setEnabled(False)
        self.btn_validar.setEnabled(False)
        self.btn_importar.setEnabled(False)
        self.btn_fechar.setEnabled(False)
        
        # Mostrar progresso
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(0)
        
        # Executar em thread
        self.import_thread = ImportThread(self.importer, self.linhas_validas)
        self.import_thread.finished.connect(self._importacao_concluida)
        self.import_thread.start()
    
    def _atualizar_progresso(self, atual, total):
        """Atualiza barra de progresso"""
        self.progress_bar.setValue(atual)
    
    def _atualizar_status(self, status):
        """Atualiza label de status"""
        self.lbl_status.setText(status)
    
    def _importacao_concluida(self):
        """Callback quando importação termina"""
        result: ImportResult = self.import_thread.result
        
        # Reabilitar botões
        self.btn_selecionar.setEnabled(True)
        self.btn_validar.setEnabled(True)
        self.btn_fechar.setEnabled(True)
        
        # Ocultar progresso
        self.progress_bar.setVisible(False)
        
        # Registrar auditoria
        if self.current_user:
            registrar_auditoria(
                self.db_manager,
                self.current_user,
                "patrimonios",
                None,
                "IMPORT",
                f"Importação em massa: {result.importados} patrimônio(s)"
            )
        
        # Mostrar resultado
        if result.sucesso:
            msg = f"✓ Importação concluída com sucesso!\n\n"
            msg += f"Patrimônios importados: {result.importados}\n"
            msg += f"Fornecedores criados: {result.fornecedores_criados}\n"
            msg += f"Setores criados: {result.setores_criados}\n"
            msg += f"Categorias criadas: {result.categorias_criadas}"
            
            QMessageBox.information(self, "Importação Concluída", msg)
            self.accept()  # Fechar diálogo
        
        else:
            msg = f"⚠ Importação concluída com erros.\n\n"
            msg += f"Patrimônios importados: {result.importados} de {result.total_linhas}\n"
            msg += f"Erros: {len(result.erros)}\n\n"
            msg += "Detalhes dos erros:\n"
            msg += "\n".join(result.erros[:10])  # Primeiros 10 erros
            
            if len(result.erros) > 10:
                msg += f"\n\n... e mais {len(result.erros) - 10} erro(s)."
            
            QMessageBox.warning(self, "Importação com Erros", msg)
